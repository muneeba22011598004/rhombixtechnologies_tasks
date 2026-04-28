import sys
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows

# ==============================
# 📁 LOAD DATA
# ==============================
df = pd.read_csv("final.csv")


# BASIC STATISTICS

numeric_df = df.select_dtypes(include='number')

summary = pd.DataFrame({
    "Mean": numeric_df.mean(),
    "Median": numeric_df.median(),
    "Min": numeric_df.min(),
    "Max": numeric_df.max()
})

# CORRELATION

correlation = numeric_df.corr()

# OUTLIERS (IQR METHOD)

outlier_data = []

for col in numeric_df.columns:
    Q1 = numeric_df[col].quantile(0.25)
    Q3 = numeric_df[col].quantile(0.75)
    IQR = Q3 - Q1

    lower = Q1 - 1.5 * IQR
    upper = Q3 + 1.5 * IQR

    outliers = df[(numeric_df[col] < lower) | (numeric_df[col] > upper)]

    for val in outliers[col]:
        outlier_data.append([col, val])

outliers_df = pd.DataFrame(outlier_data, columns=["Column", "Outlier_Value"])


# GRAPHS SAVE KARNA


# Histogram
plt.figure()
plt.hist(df["Actual_gross"])
plt.title("Histogram of Actual Gross")
plt.xlabel("Actual Gross")
plt.ylabel("Frequency")
plt.savefig("hist.png")
plt.close()

# Bar Chart
top5 = df.nlargest(5, "Actual_gross")
plt.figure()
plt.bar(top5["Tour_title"], top5["Actual_gross"])
plt.xticks(rotation=45)
plt.title("Top 5 Tours")
plt.savefig("bar.png")
plt.close()

# Line Graph
plt.figure()
plt.plot(df["start_year"], df["Actual_gross"])
plt.title("Gross over Years")
plt.xlabel("Year")
plt.ylabel("Gross")
plt.savefig("line.png")
plt.close()

# EXCEL FILE CREATE

wb = Workbook()

# Sheet 1: Data
ws1 = wb.active
ws1.title = "Data"

for row in dataframe_to_rows(df, index=False, header=True):
    ws1.append(row)

# Sheet 2: Statistics
ws2 = wb.create_sheet("Statistics")

for row in dataframe_to_rows(summary, index=True, header=True):
    ws2.append(row)

# Sheet 3: Correlation
ws3 = wb.create_sheet("Correlation")

for row in dataframe_to_rows(correlation, index=True, header=True):
    ws3.append(row)

# Sheet 4: Outliers
ws4 = wb.create_sheet("Outliers")

for row in dataframe_to_rows(outliers_df, index=False, header=True):
    ws4.append(row)

# Sheet 5: Graphs
ws5 = wb.create_sheet("Graphs")

ws5.add_image(XLImage("hist.png"), "A1")
ws5.add_image(XLImage("bar.png"), "A20")
ws5.add_image(XLImage("line.png"), "A40")

# Save Excel
wb.save("final_analysis.xlsx")

print("DONE! Excel file created: final_analysis.xlsx")
